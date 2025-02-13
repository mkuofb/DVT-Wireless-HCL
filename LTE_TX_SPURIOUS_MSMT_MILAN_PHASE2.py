#!/usr/bin/env python3
import sys
import os
import time
import re
import shutil
# from openpyxl import Workbook
import openpyxl
import datetime
import json
import csv
import logging
import socket
import xlrd
# import pyvisa_py
import pyvisa
from RsInstrument import *
from RsInstrument.RsInstrument import RsInstrument, BinIntFormat
import numpy as np
import itertools
import pandas as pd
from operator import itemgetter
from ppadb.client import Client as AdbClient
from glob import glob
from pandas import ExcelWriter
from ctf_json_data import CtfJsonData


# Setup Configuration
test_environment = "phase2"  # phase1 or phase2
cmw_interface = "LAN"
fsw_interface = "LAN"
sw_box_interface = "GPIB"
# pwr_supp_interface = "GPIB"
dut_adb_serial_number = "225c2b92"
screenshot_host_ip = "192.168.0.219"
swbox = None
pathloss_file = {1:'S46_B1-C1_pathloss', 2:'S46_B2-C2_pathloss', 3:'S46_B3-C3_pathloss',
                 4:'S46_B4-C4_pathloss', 5:'S46_B5-C5_pathloss', 6:'S46_B6-C6_pathloss',
                 0:'dummy_FSW_pathloss'}

# Config Phase1 or Phase2 settings
if test_environment.upper() == "PHASE1":
    cmw_interface = "LAN"
    fsw_interface = "LAN"
    sw_box_interface = "NA"
    dut_adb_serial_number = "225c2b92"
    screenshot_host_ip = "192.168.0.219"
else:
    cmw_interface = "LAN"
    fsw_interface = "GPIB"
    sw_box_interface = "GPIB"
    dut_adb_serial_number = "225c2b92"  # "4D123456789"
    screenshot_host_ip = "192.168.0.219"  # "172.22.1.5"

# create a log
logger = logging.getLogger(__name__)
# output format
logger_format = "%(relativeCreated)8d %(levelname)-12s %(filename)-35s %(funcName)25s %(lineno)5s: %(message)s"
fsw_host = 0
fsw_socket_port = 0
# get a logging handle to the screen
sh = logging.StreamHandler()
# format the output
sh.setFormatter(logging.Formatter(logger_format))
logger.addHandler(sh)
logger.setLevel(logging.DEBUG)
# Making a directory For the Run
mydir = os.path.join('{0}'.format(os.getcwd().split('D')[0]), "LTE_TX_SPURIOUS_MSMT",
                     '{0}_{1}_{2}'.format("FULL_RUN_LOGS", datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S'), sys.argv[1],
                                          os.getcwd().split('D')[0]))
os.makedirs(mydir)
mydir_d = mydir.replace('\\', '\\\\')
logger.debug("Result log folder created succesfully {0}".format(mydir_d))


# while os.path.exists("{0}\\stability_run\\LTE_TX_Band13_max".format(os.getcwd().split('D')[0])) ==
dest = os.path.join('{0}'.format(os.getcwd().split('D')[0]), "LTE_TX_SPURIOUS_MSMT_CTF",
                    'LTE_TX_{0}_{1}'.format(sys.argv[1], "FULL_RUN_LOGS", os.getcwd().split('D')[0]))
# os.makedirs(dest)
dest_d = dest.replace('\\', '\\\\')
logger.debug("OUT put logs will be stored in {0}".format(dest_d))
# Screenshot folder  under LTE_TX_SPURIOUS_MSMT_CTF\LTE_TX_harmonics|Full runlogs)

logfilename = dest_d.replace('.py', '').replace('.PY', '')
timestr_l = time.strftime("%Y%m%d-%H%M%S")
logfilename = logfilename + '{0}.log'.format(timestr_l)
#logfilename = mydir_d +'{0}.log'.format(timestr_l)
#logfilename = dest +'{0}.log'.format(timestr_l)
fh = logging.FileHandler(logfilename, mode='w')
# format outout
fh.setFormatter(logging.Formatter(logger_format))
# add file handle to logger instance
logger.addHandler(fh)


# Menu for input file
while True:
    logger.debug("choose the input file")
    logger.debug("1.Harmonics_Sweep_v1.xlsx")
    logger.debug("2.Harmonics_Sweep_v3.xlsx")
    logger.debug("3.Harmonics_Sweep_v4.xlsx")
    logger.debug("4.Harmonics_Sweep_v5.xlsx")
    logger.debug("5.Harmonics_Sweep_v6.xlsx")
    choice= int(input("enter your choice"))
    if choice == 1:
        LTE_FILE = 'Harmonics_Sweep_v1.xlsx'
        break
    elif choice == 2:
        LTE_FILE = 'Harmonics_Sweep_v3.xlsx'
        break
    elif choice == 3:
        LTE_FILE = 'Harmonics_Sweep_v4.xlsx'
        break
    elif choice == 4:
        LTE_FILE = 'Harmonics_Sweep_v5.xlsx'
        break
    elif choice == 4:
        LTE_FILE = 'Harmonics_Sweep_v6.xlsx'
        break
    else:
        logger.debug("No file chosen")

# *************************************Create Test File From Raw file **********************

header_list = pd.read_excel(LTE_FILE, skiprows=3).columns

Band_filter = pd.read_excel(LTE_FILE, skiprows=1).columns
Port_filter = pd.read_excel(LTE_FILE, skiprows=2).columns
logger.debug(Band_filter[2])# 2
logger.debug(Port_filter[2])


if os.path.exists("C:\Test\DVT-Wireless-HCL\Harmonics_Sweep_v2.xlsx"):
    os.remove("Harmonics_Sweep_v2.xlsx")
else:
    logger.debug("file is not found")
file_location = "C:\Test\DVT-Wireless-HCL\Harmonics_Sweep_v2.xlsx"
df = pd.DataFrame(columns=header_list)
writer = ExcelWriter(file_location)
df.to_excel(writer, 'Harmonics_Sweep_v1', index=0, startrow=1)
writer.save()  # this create Excel file with your desired titles

# **********************************************************
# total_bands = ['Band2', 'Band4', 'Band5', 'Band12', 'Band13', 'Band66']
# to pick min/max power input file from ctf command ine
# *************************************COpy DATA From Raw file to Test FIle**********************
# Is use to create a reference of the Excel to wb
wb1 = openpyxl.load_workbook(LTE_FILE)
wb2 = openpyxl.load_workbook('Harmonics_Sweep_v2.xlsx')

# sband = [3, 2]
# Refrence the workbook to the worksheets
sh1 = wb1["Harmonics_Sweep_v1"]
sh2 = wb2["Harmonics_Sweep_v1"]  # use same sheet name, different workbook
# print(sh1.iter_rows)#


# list = cell value( 1,2)
# r = [2,4,12]
for row in sh1.iter_rows():
    # print(column[0].value)

    if row[7].value == 2 or row[7].value == 4 or row[7].value == 5 or row[7].value == 12 or row[7].value == 13 or row[7].value == 66:  # filter on first column with value 16 [band == 2]
        # print(row[0].value)
     #if row[7].value ==13:
        if row[3].value == Port_filter[2] :
            sh2.append((cell.value for cell in row))

wb1.save(LTE_FILE)
wb2.save("Harmonics_Sweep_v2.xlsx")
# *************************************COpy DATA From Raw file to Test FIle**********************

if sys.argv[1] == 'Harmonics' or sys.argv[1] == 'harmonics' or sys.argv[1] == None:
    lte_tx_input_file = 'Harmonics_Sweep_v2.xlsx'
    logger.debug("TEST is based on 3gpp MAX power Value -23dbm")
else:
    logger.debug(" Please check the command line argument  either MAX/max or MIN/min valid, Refer to Test Description")

# pick the min/max input file from current directory
inp_file = ("{1}\\{0}".format(lte_tx_input_file, os.getcwd()))
logger.debug("input file is {0}".format(inp_file))
wb = xlrd.open_workbook(inp_file)
#sheet = wb.get_sheet_by_name('Harmonics_Sweep_v1') 
sheet = wb.sheet_by_index(0)
#max_rows = sheet.max_rows
#logger.debug(" max rows {0}".format(max_rows))
# print(sheet)
# sheet = wb.sheet_by_name("Harmonics_Sweep_v1")
"""
logic to convert the raw (harmonic_file.excel )>> Milan .excel
LTE SPUR == MIlan.excel
LTE_file = Milan.excel
inp_file = ("{1}\\{0}".format(LTE_FILE, os.getcwd()))
wb_milan = xlrd.open_workbook(inp_file)
sheet = wb_milan.sheet_by_index(0)
"""
"""
for i in total_bands:
    if i == sys.argv[1]:
        #sheet = wb.sheet_by_index(i)
        sheet = wb.sheet_by_name(sys.argv[1]) # band
    else:
        logger.debug(" not valid argument to select the input band sheet")
"""
"""
header_list = pd.read_excel('Harmonics_Sweep_v1.xlsx', skiprows=3).columns
#print(header_list)
#os.remove("/tmp/<file_name>.txt")

file_location = "C:\Test\DVT-Wireless-HCL\Harmonics_Sweep_v2.xlsx"
#if
df = pd.DataFrame(columns=header_list)
writer = ExcelWriter(file_location)
df.to_excel(writer, 'Harmonics_Sweep_v1', index=0, startrow=1)
writer.save()  # this create Excel file with your desired titles
"""

# row_count = int(sys.argv[2]) # start variant
# start of the while  loop
row_count = 8
end_of_loop = 9
loop_rst = 1
while int(row_count) < end_of_loop:  # stop variant
    pos = 0
    ##Loop startes for
    # time.sleep(120)
    logger.debug(132 * '_')
    # logger.debug(type(row_count))
    # logger.debug(type(sheet.celTEST_BWl_value(row_count, 0)))
    # band_in = sheet.cell_value(row_count, 0)
    # TEST_BAND = int(sheet.cell_value(row_count, 0))
    # pick_band = 2 int(sheet.cell_value(row_count, 0) x
    # support_list = [2,4,5,12,13,66]
    # for i in suppor
    Temp = str(sheet.cell_value(row_count, 0))
    Voltage = float(sheet.cell_value(row_count, 1))
    Tech = str(sheet.cell_value(row_count, 2))
    Antena = str(sheet.cell_value(row_count, 3))
    Call_Type = str(sheet.cell_value(row_count, 4))
    NS_Filter = int(sheet.cell_value(row_count, 5))

    pmx= int(sheet.cell_value(row_count, 6))
    logger.debug(pmx)
    logger.debug(type(pmx))
    #pmx = 24
    Band = int(sheet.cell_value(row_count, 7))
    logger.debug(Band)
    logger.debug(type(Band))
    """
    support_list = [2, 4, 5, 12, 13, 66]
    for i in support_list:
        if Band != i:
            break
    else:
        print('-- Finish inner loop without BREAK')
        continue

   """
    UL_EARFCN = int(sheet.cell_value(row_count, 8))
    DL_EARFCN = int(sheet.cell_value(row_count, 9))
    Channel_Bandwidth = (sheet.cell_value(row_count, 10))

    UL_RB_Allocation = int(sheet.cell_value(row_count, 11))

    logger.debug("value of UL_RB_Allocation is{0} ".format(UL_RB_Allocation))
    UL_RB_Start = int(sheet.cell_value(row_count, 12))
    logger.debug(type(UL_RB_Start))
    logger.debug("value of rb is{0} ".format(UL_RB_Start))
    UL_MODULATION = str(sheet.cell_value(row_count, 13))
    logger.debug("value of rb is{0} ".format(UL_MODULATION))
    UL_TBS = int(sheet.cell_value(row_count, 14))+4
    logger.debug("value of rb is{0}".format(UL_TBS))
    DL_RB_Allocation = int(sheet.cell_value(row_count, 15))
    DL_RB_Start = int(sheet.cell_value(row_count, 16))
    DL_MODULATION = str(sheet.cell_value(row_count, 17))
    DL_TBS = int(sheet.cell_value(row_count, 18))

    freq_start_r1 = int(sheet.cell_value(row_count, 19))
    freq_stop_r1 = int(sheet.cell_value(row_count, 20))
    DelFOB = int(sheet.cell_value(row_count, 21))
    switch_path = int(sheet.cell_value(row_count, 22))
    ref_level = int(sheet.cell_value(row_count, 23))
    attn = int(sheet.cell_value(row_count, 24))
    rbw = float(sheet.cell_value(row_count, 25))
    vbw = float(sheet.cell_value(row_count, 26))
    pre_amp = str(sheet.cell_value(row_count, 27))
    Trace = str(sheet.cell_value(row_count, 28))
    Detector = str(sheet.cell_value(row_count, 29))
    sweep_time = (sheet.cell_value(row_count, 30))
    sweep_point = int(sheet.cell_value(row_count, 31))
    sweep_count = int(sheet.cell_value(row_count, 32))
    spec_limit = int(sheet.cell_value(row_count, 33))
    UL_ATT = (sheet.cell_value(row_count, 34))
    DL_ATT = (sheet.cell_value(row_count, 35))

    # set log level
    # logger.setLevel(logging.DEBUG)
    logger.debug("Band: 	\t Bandwidth : (MHz) \t	DL Frequency:  (MHz)	\tDL RB Allocation: 	\t DL RB Start: 	\t float(Power Level)")
    logger.debug(" {0}	 \t\t  :{1} (MHz) \t\t	: {2} (MHz)\t	\t\t: {3}	   \t\t\t : {4}	\t\t\t  {5}  ".format(Band,Channel_Bandwidth,DL_EARFCN, DL_RB_Allocation, DL_RB_Start,pmx))

    logger.debug(132 * '_')


    # socket class
    class c_socket(object):

        def __init__(self, host, port, **kwargs):
            """ class Socket

                :param host: str, - CMW IP address or Host Name
                :param port: int, - CMW socket port - default 5025 for SCPI
                :param timeout: float = 3, - timeout in seconds
                :param termchar: str = '\n' -  Delimiter character

            Example:
                kwargs = {'timeout': 60, 'termchar': '\n'}
                mySocket = Socket('172.24.224.111', 5025)
            """            # get a log instance from logging
            if 'logger' in kwargs:
                self._logger = logging.getlogger('logger')
            else:
                self._logger = logging.getLogger(__name__)

            # size of the incoming buffer 4096 bytes
            self._buffer_size = 4096
            # pompt decorator for the logging
            self._logindicator_write = '>>'
            self._logindicator_read = '<<'
            # boolean
            self.is_connected = False
            # connection type
            self.type = 'socket'
            # populate the class
            # host or ip address
            if 'host' in kwargs:
                self.host = kwargs['host']
            else:
                self.host = host

            # socket port
            if 'port' in kwargs:
                self.port = kwargs['port']
            else:
                self.port = port

            # timeout
            if 'timeout' in kwargs:
                self.__timeout = kwargs['timeout']
            else:
                self.__timeout = 6

            # incoming byte delimiter
            if 'termchar' in kwargs:
                self.termchar = kwargs['termchar']
            else:
                self.termchar = '\n'

            # socket instance
            self._sock = None
            self.ask = self.query

        def __del__(self):
            """ __del__ - Class destructor
            """

            if self._sock is not None:
                # close socket if instance is not None
                self._close()
            self._sock = None

        def close(self):
            """ _close() - Close socket connection
            """
            try:
                if self._sock is not None:
                    self._logger.debug('Close Socket Handle: {0}'.format(self._sock))
                    self._close()
                self.is_connected = False

            except Exception as Err:
                raise Err

        def connect(self):
            """ connect() - Attempt to open a socket connection
            """
            try:

                if self.is_connected:
                    self.close()
                    self._sock = None

                # get a socket instance
                self._sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

                # configure the socket interface
                self._sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_NODELAY, 1)

                # set socket timeout
                self.timeout = self.__timeout

                self._write = self._sock.sendall
                self._read = self._sock.recv
                self._close = self._sock.close

                # attempt to connect - if no connection then assume running in debug mode so
                # SCPI commands can be logged
                self._logger.debug('Open SOCKET Connection @: {0}:{1:d}'.format(self.host, self.port))
                try:
                    self._debug_mode = False
                    self._sock.connect((self.host, self.port))
                    self.timeout = self.__timeout

                except:
                    self._logger.error("SCPI Connection failed - run debug mode only ")
                    self._debug_mode = True

            except socket.error as Err:
                raise

            except Exception as Err:
                msg = 'Could not connect to host {0}:{1}'.format(self.host, self.port)
                self._logger.exception(msg)
                raise ValueError(msg)

            self.is_connected = True
            self._logger.debug(
                'SOCKET Connection Successfully Open with: {0}:{1} - SOCKET Handle: {2}'.format(self.host, self.port,
                                                                                                [self._sock]))

        def _check_connection(self):
            """ _check_connection() - Check socket connection

                :return: bool
            """
            if not self.is_connected:
                msg = 'Not connected!'
                self._logger.exception(msg)
                raise ValueError(msg)

        def _reconnect(self):
            """ _reconnect() - Re-connect a socket connection
            """
            self.close()
            self.connect()

        def write(self, scpi_cmd):
            """ write(scpi_cmd) - send SCPI command to socket interface

                :param scpi_cmd:

            """
            try:
                self._check_connection()

                self._logger.debug('{0}:{1} >> {2}'.format(self.host, self.port, scpi_cmd))

                command = (scpi_cmd + self.termchar).encode()
                self._sock.sendall(command)

            except Exception as Err:
                raise Err

        def read(self):
            """ read() - read incoming bytes from socket interface

                :return:  str
            """
            try:

                buffin = ''

                if self._debug_mode:
                    return ['']

                start = time.time()
                while True:

                    response = b''
                    self._check_connection()

                    response = self._sock.recv(4096)
                    buffin = buffin + response.decode()

                    if self.termchar in buffin:
                        buffin = buffin.replace(self.termchar, '')
                        break

                    if time.time() - start > self.timeout:
                        self._logger.warning('Timeout while reading socket.')
                        return ''

                if (buffin[0] == '"' and buffin[-1] == '"') or (
                        buffin[0] == '"' and buffin[-1] == '"' and ';' in buffin):
                    buffout = [buffin]
                elif ',' in buffin:
                    buffout = buffin.split(',')
                else:
                    buffout = [buffin]

                self._logger.debug('values are {0}:{1} << [{2}]'.format(self.host, self.port, ','.join(buffout)))

                return buffout

            except Exception as Err:
                raise Err

        def _read_until(self):
            """ _read_until(()

                :return: response_byte - str byte
            """

            response_byte = b''
            start = time.time()
            while time.time() - start < self.timeout:

                try:
                    # read bytes from socket -  default buffer size 4096 bytes
                    response_temp = self._read(self._buffer_size)
                except:
                    self._logger.warning('Timeout while reading socket.')
                    self._reconnect()
                    return response_byte

                # add read bytes to the response
                response_byte += response_temp
                if self.termchar.encode('utf-8') in response_byte:
                    # exit if delimiter char found in the response_byte
                    return response_byte[:-1]

            self._logger.warning('Timeout waiting on termchar.')
            self._reconnect()
            return response_byte

        def query(self, command):
            """ query(command):  Send a SCPI command and retrieve the answer

                :param command: - str
                :return:  str
            """

            self._check_connection()
            self.write(command)
            response = ['']
            if '?' in command:
                response = self.read()
            return response

        # property timeout
        def __set_timeout(self, timeout_s):

            if timeout_s != self.__timeout:
                self.__timeout = timeout_s

            if self._sock is not None:
                self._sock.settimeout(self.__timeout)

        def __get_timeout(self):

            if self._sock is not None:
                self.__timeout = self._sock.gettimeout()
            if self.__timeout is None:
                self.__timeout = 6
            return self.__timeout

        timeout = property(__get_timeout, __set_timeout, None, """ timeout """)


    class c_gpib():
        """ GPIB Interface """
        def __init__(self, instr_name):
            self.cmw_gpib_addr = 'GPIB0::20::INSTR'
            self.fsw_gpib_addr = 'GPIB0::21::INSTR'
            self.sw_box_gpib_addr = 'GPIB0::7::INSTR'  # 'GPIB0::4::INSTR'
            self.pwr_supply_gpib_addr = 'GPIB0::5::INSTR'
            self.instr_name = instr_name
            self.instr = None

        def connect(self):
            try:
                rm = pyvisa.ResourceManager()
                if self.instr_name.upper() == 'CMW':
                    self.instr = rm.open_resource(self.cmw_gpib_addr)
                elif self.instr_name.upper() == 'FSW':
                    self.instr = rm.open_resource(self.fsw_gpib_addr)
                elif self.instr_name.upper() == 'SW_BOX':
                    self.instr = rm.open_resource(self.sw_box_gpib_addr)
                else:
                    logger.error("Invalid instrument name!")
                logger.info("Connect {} GPIB interface".format(self.instr_name.upper()))
            except Exception as Err:
                raise Err

        def write(self, cmd):
            self.instr.write(cmd)

        def read(self):
            self.instr.read()

        def ask(self, cmd):
            return self.instr.query(cmd)

        def close(self):
            self.instr.close()


    class c_switchbox():
        """ KEITHLEY S46 Switch Box Control """
        def __init__(self):
            self.swbox = None

        def connect(self):
            self.swbox = c_gpib("SW_BOX")
            self.swbox.connect()

        def write(self, cmd):
            self.swbox.write(cmd)

        def ask(self, cmd) -> str:
            msg = self.swbox.query(cmd)
            return msg

        def read(self) -> str:
            msg = self.swbox.read()
            return msg

        def query_closed_channels(self):
            chs = self.ask("ROUT:CLOS?")
            logger.info("SW Box closed channels: {}".format(chs))

        def set_all_port_open(self):
            logger.info("SW Box open all ports")
            cmd = "ROUT:OPEN:ALL"
            self.write(cmd)

        def set_wire_only_path(self):
            logger.info("SW Box switch to wire path")
            cmd = ["ROUT:OPEN:ALL", "ROUT:CLOS (@7,13)"]
            self.write(cmd[0])
            self.write(cmd[1])

        def set_10db_atten_path(self):
            logger.info("SW Box switch to 10dB atten path")
            cmd = ["ROUT:OPEN:ALL", "ROUT:CLOS (@8,14)"]
            self.write(cmd[0])
            self.write(cmd[1])

        def set_1p2g_hpf_path(self):
            logger.info("SW Box switch to 1.2GHz HPF path")
            cmd = ["ROUT:OPEN:ALL", "ROUT:CLOS (@9,15)"]
            self.write(cmd[0])
            self.write(cmd[1])

        def set_3g_hpf_path(self):
            logger.info("SW Box switch to 3.0GHz HPF path")
            cmd = ["ROUT:OPEN:ALL", "ROUT:CLOS (@10,16)"]
            self.write(cmd[0])
            self.write(cmd[1])

        def set_6p4g_hpf_path(self):
            logger.info("SW Box switch to 6.4GHz HPF path")
            cmd = ["ROUT:OPEN:ALL", "ROUT:CLOS (@11,17)"]
            self.write(cmd[0])
            self.write(cmd[1])

        def set_18g_hpf_path(self):
            logger.info("SW Box switch to 18.0GHz HPF path")
            cmd = ["ROUT:OPEN:ALL", "ROUT:CLOS (@12,18)"]
            self.write(cmd[0])
            self.write(cmd[1])

        def set_sw_box(self, switch_path):
            logger.info("SW Box switch_path = {}".format(switch_path))

            if not (switch_path >= 1 or switch_path <= 6):
                logger.info("Invalid switch_path value. Use default value 1")
                switch_path = 1

            if switch_path == 1:
                self.set_wire_only_path()
            elif switch_path == 2:
                self.set_10db_atten_path()
            elif switch_path == 3:
                self.set_1p2g_hpf_path()
            elif switch_path == 4:
                self.set_3g_hpf_path()
            elif switch_path == 5:
                self.set_6p4g_hpf_path()
            elif switch_path == 6:
                self.set_18g_hpf_path()


    class c_lte(object):

        def __init__(self):
            self.sua = 1
            self.rf_port = "RF1C"
            self.converter = 1
            self.qci = 5
            self.rb_d = {1.4: 6, 3: 15, 5: 25, 10: 50, 15: 75, 20: 100}
            self.bw_d = {2: [1.4, 3, 5, 10, 15, 20],
                         4: [1.4, 3, 5, 10, 15, 20],
                         5: [1.4, 3, 5, 10],
                         12: [1.4, 3, 5, 10],
                         13: [5, 10],
                         66: [1.4, 5, 20], }
            self.band_d = {2: [600, 600, 1930.0, 80.0],
                           4: [1950, 450, 2110.0, 400.0],
                           5: [2400, 250, 869.0, 45.0],
                           12: [5010, 170, 729.0, 30.0],
                           13: [5180, 100, 746.0, -31.0],
                           66: [66436, 700, 2110, 400]
                           }
            self.low_prb = 0
            self.rb_max = {1.4: 5, 5: 24, 10: 49, 20: 99}
            self.suw = 1
            self.suu = 1
            self.sua = 1
            self.rf_port = "RF1C"
            self.converter = 1
            self.opc_key = "00000000000000000000000000000000"
            self.secret_key = "000102030405060708090A0B0C0D0E0F"
            self.milenage = False
            self.band = Band
            self.tx_close_loop_power = pmx


    rb = c_lte().rb_d.values()
    tx_close_loop_target_power = c_lte().tx_close_loop_power
    srb = c_lte().rb_max.values()
    logger.debug(list(srb))

    # Look for UE device
    # TODO: make CLI argument

    # Default is "127.0.0.1" and 5037
    client = AdbClient(host="127.0.0.1", port=5037)
    logger.info('version of module : %d', client.version())

    # get list of devices
    devices = client.devices()

    device_found = False
    for dev in devices:
        logger.info('detected : %s', dev.serial)
        if dev.serial == dut_adb_serial_number:
            device_found = True
            break

    if not device_found:
        logger.error('could not find %s. terminating script', dut_adb_serial_number)
        sys.exit(0)

    # connect to selected device
    device = client.device(dut_adb_serial_number)
    logger.info('connected to adb device %s', dut_adb_serial_number)

    # CMW instance
    cmw = None
    cmw_host = '127.0.0.1'
    cmw_socket_port = 5025
    cmw_socket_timeout = 6
    cmw_socket_termchar = '\n'
    fsw_socket_timeout = 6
    fsw_socket_termchar = '\n'

    DAU_PRESENT = True
    IMS_FLAG = True

    REGISTER_TIMEOUT = 60

    dau_service_l = ['DNS', 'FTP', 'HTTP', 'IMS2', 'EPDG']

    lte = c_lte()

    epdg_ip_v4 = "192.168.1.201"
    epdg_ip_v6 = "fc01::1"
    dau_ipv4_config = 'internal'
    dau_ipv6_config = 'internal'

    try:

        # read json file
        data = None
        f_json = os.path.curdir + os.sep + 'config_input.json'
        # fsw_json = os.path.curdir + os.sep + 'config.json'
        if os.path.exists(f_json):
            with open(f_json) as json_file:
                data = json.load(json_file)

            if 'cmw' in data:
                if 'host' in data['cmw']:
                    cmw_host = data['cmw']['host']

                if 'socket_port' in data['cmw']:
                    cmw_socket_port = data['cmw']['socket_port']

                if 'socket_timeout' in data['cmw']:
                    cmw_socket_timeout = data['cmw']['socket_timeout']

                if 'socket_termchar' in data['cmw']:
                    cmw_socket_termchar = data['cmw']['socket_termchar']

            if 'fsw' in data:
                if 'host' in data['fsw']:
                    fsw_host = data['fsw']['host']

                if 'socket_port' in data['fsw']:
                    fsw_socket_port = data['fsw']['socket_port']

                if 'socket_timeout' in data['fsw']:
                    fsw_socket_timeout = data['fsw']['socket_timeout']

                if 'socket_termchar' in data['fsw']:
                    fsw_socket_termchar = data['fsw']['socket_termchar']
            """
            if fsw_param' in data :
                rbw_value = data['fsw_param']['rbw'] # [ 1, 10 ,20 ,30 ]
                logger.
            """
            # Reading FSw spurious msmt settings from config_r.json
            if 'fsw_params' in data:
                """
                if 'sweep_points' in data['fsw_params']:
                    sweep_point_value = data['fsw_params']['sweep_points'] # [ 800, 5000, 3300, 34000 ]
                if 'ranges' in data['fsw_params']:
                    range_num = data['fsw_params']['ranges'] # [ 800, 5000, 3300, 34000 ]
                if 'ref_lev' in data['fsw_params']:
                    ref_lev = data['fsw_params']['ref_lev']
                if 'att_rf' in data['fsw_params']:
                    att_rf = data['fsw_params']['att_rf']
                if 'transducer' in data['fsw_params']:
                    transducer = data['fsw_params']['transducer']

                if 'freq_start' in data['fsw_params']:
                    freq_start = data['fsw_params']['freq_start']
                if 'freq_stop' in data['fsw_params']:
                    freq_stop = data['fsw_params']['freq_stop']

                if 'rbw_in' in data['fsw_params']:
                    rbw_in = data['fsw_params']['rbw_in']
                if 'rbw_high' in data['fsw_params']:
                    rbw_high = data['fsw_params']['rbw_high']
                if 'vbw_in' in data['fsw_params']:
                    vbw_in = data['fsw_params']['vbw_in']
                if 'detector_type' in data['fsw_params']:
                    detector_type = data['fsw_params']['detector_type']
                if 'limit' in data['fsw_params']:
                    limit = data['fsw_params']['limit']
                if 'sweep_time' in data['fsw_params']:
                    sweep_time = data['fsw_params']['sweep_time']
                """
            if 'lte' in data:

                if 'sua' in data['lte']:
                    lte.sua = data['lte']['sua']

                if 'rf_port' in data['lte']:
                    lte.rf_port = data['lte']['rf_port']

                if 'converter' in data['lte']:
                    lte.converter = data['lte']['converter']

                if 'band' in data['lte']:
                    lte.band = data['lte']['band']

                if 'bw' in data['lte']:
                    lte.bw = data['lte']['bw']

                if 'rs_epre' in data['lte']:
                    lte.rs_epre = data['lte']['rs_epre']

                if 'imsi' in data['lte']:
                    lte.imsi = data['lte']['imsi']

                if 'apn' in data['lte']:
                    lte.apn = data['lte']['apn']

                if 'qci' in data['lte']:
                    lte.qci = data['lte']['qci']

        MCC = lte.imsi[0:3]
        MNC = lte.imsi[3:6]
        IMSI = lte.imsi[6:]
        SIM = 'TMO'

        # get argument parsed
        if '--cmw' in sys.argv:
            cmw_host = sys.argv[sys.argv.index("--cmw") + 1]

        if '--lte_band' in sys.argv:
            logger.debug(" inside sys arg")
            Band = int(sys.argv[sys.argv.index("--lte_band") + 1])
            logger.debug(Band)

        if '--lte_bw' in sys.argv:
            Channel_Bandwidth = float(sys.argv[sys.argv.index("--Channel_Bandwidth") + 1])

        if '--lte_dl_chan' in sys.argv:
            DL_EARFCN = int(sys.argv[sys.argv.index("--DL_EARFCN") + 1])

        if '--lte_tx_close_loop_power' in sys.argv:
            tx_close_loop_target_power = float(sys.argv[sys.argv.index("--tx_close_loop_target_power") + 1])

        if '--dau_ipv4_external' in sys.argv:
            dau_ipv4_config = 'external'

        if '--dau_ipv6_external' in sys.argv:
            dau_ipv6_config = 'external'

        logger.info('{0}'.format(132 * '-'))
        logger.info('***** Script Setup CMW for LTE call *****')
        logger.info('{0:<20} : {1}'.format('Date', time.asctime()))
        logger.info('{0:<20} : {1}'.format('Computer', socket.gethostname().upper()))
        logger.info('{0:<20} : {1}'.format('Operating System', os.name.upper()))
        logger.info('{0:<20} : {1}'.format('Target OS', sys.platform.upper()))
        logger.info('{0:<20} : {1}'.format('Python Version', sys.version.upper()))
        logger.info('{0:<20} : {1}'.format('Python Script', sys.argv[0]))
        logger.info('{0}'.format(132 * '-'))

        logger.info('{0}'.format(132 * '-'))
        logger.info('Get working directory')
        workdir = os.getcwd()
        logger.info('{0}'.format(workdir))
        logger.info('{0}'.format(132 * '-'))
        # time.sleep(10)

        # Init Switch Box and open all paths.
        if sw_box_interface == "GPIB":
            swbox = c_switchbox()
            swbox.connect()
            swbox.set_all_port_open()

        # Init CMW
        if 1:
            logger.info('{0}'.format(132 * '-'))
            logger.info('Create a Socket Class object to remote the CMW')
            kwargs = {'timeout': cmw_socket_timeout, 'temchar': cmw_socket_termchar}

            # get instance for controlling CMW
            if cmw_interface == "LAN":
                cmw = c_socket(cmw_host, cmw_socket_port, **kwargs)
            else:
                cmw = c_gpib("CMW")

            # connect to CMW
            cmw.connect()

            # set CMW display
            logger.info('{0}'.format(132 * '-'))

            cmw.write('SYST:DISP:UPD ON')

            logger.info('{0}'.format(132 * '-'))
            logger.info('RESET CMW')
            cmw.timeout = 1080.0
            cmw.write("*CLS")
            cmw.write("SYST:PRES:ALL")
            cmw.ask("*OPC?")
            time.sleep(0.5)

            logger.info('{0}'.format(132 * '-'))
            logger.info('QUERY ID')

            if not hasattr(cmw, 'id'):
                setattr(cmw, 'id', 'CMW')
            if not hasattr(cmw, 'sn'):
                setattr(cmw, 'sn', '000000')
            if not hasattr(cmw, 'BASE_FW'):
                setattr(cmw, 'BASE_FW', 'CMW')

            buffin = ''
            if cmw_interface == "LAN":
                buffin = cmw.ask("*IDN?")
            else:
                buffin = cmw.ask("*IDN?").split(',')

            logger.debug(buffin)
            if 'k50' in buffin[2]:
                cmw.id = 'CMW500'
            elif 'k55' in buffin[2]:
                cmw.id = 'CMWC'
            elif 'k29' in buffin[2]:
                cmw.id = 'CMW290'

            # get serial number
            cmw.sn = buffin[2].split('/')[1]

            # Get BASE FW revision
            cmw.BASE_FW = buffin[3]

            # get cmw options
            logger.info('{0}'.format(132 * '-'))
            logger.info('QUERY CMW Option')
            if not hasattr(cmw, 'opt'):
                setattr(cmw, 'opt', [])
            cmw.opt = cmw.ask("*OPT?")
            time.sleep(3)

            if not hasattr(cmw, 'hw_list'):
                setattr(cmw, 'hw_list', [])

            if cmw_interface == "LAN":
                cmw.hw_list = cmw.ask('SYST:BASE:OPT:LIST? HWOP,FUNC')[0][1:-1].split(',')
            else:
                cmw.hw_list = cmw.ask('SYST:BASE:OPT:LIST? HWOP,FUNC')[1:-1].split(',')

            if not hasattr(cmw, 'sw_list'):
                setattr(cmw, 'sw_list', [])

            if cmw_interface == "LAN":
                cmw.sw_list = cmw.ask('SYST:BASE:OPT:LIST? SWOP,VALid')[0][1:-1].split(',')
            else:
                cmw.sw_list = cmw.ask('SYST:BASE:OPT:LIST? SWOP,VALid')[1:-1].split(',')

            # get CMW LTE FW revision
            logger.info('{0}'.format(132 * '-'))
            logger.info('QUERY LTE FW Revision')
            if not hasattr(cmw, 'lte_sign_fw'):
                setattr(cmw, 'lte_sign_fw', [])
            cmw.lte_sign_fw = cmw.ask("SYST:OPT:VERS? \"CMW_LTE_Sig\"")[0][1:-1]

            if not hasattr(cmw, 'lte_mev_fw'):
                setattr(cmw, 'lte_mev_fw', [])
            cmw.lte_mev_fw = cmw.ask("SYST:OPT:VERS? \"CMW_LTE_Meas\"")[0][1:-1]

            # Check presence of DAU
            logger.info('{0}'.format(132 * '-'))
            if cmw.hw_list.count("H450A") != 0 or cmw.hw_list.count("H450B") != 0 or cmw.hw_list.count(
                    "H450D") != 0 or cmw.hw_list.count("H450H") != 0 or cmw.hw_list.count("H450I") != 0:
                DAU_PRESENT = True

            else:
                DAU_PRESENT = False
                logger.debug('Error: LTE call requires DAU option CMW-B450 ')
                raise ValueError('LTE call requires DAU option CMW-B450')

            logger.info('{0}'.format(132 * '-'))

            erreur = cmw.ask("SYST:ERR:ALL?")
            if erreur[0] != '0':
                raise ValueError(erreur)
                # End Init CMW
                # Configure CMW LTE CELL CMW DAU IMS2 server and VoLTE subscriber
            if 1:
                logger.info('{0}'.format(132 * '-'))
                logger.info('LTE Cell - TURN LTE CELL OFF')
                buffin = cmw.ask("SYST:SIGNaling:ALL:OFF;*OPC?")

                TIMEOUT = 45.0
                tstart = time.time()
                if cmw_interface == "LAN":
                    buffin = cmw.ask("SOUR:LTE:SIGN:CELL:STAT:ALL?")
                    while not (buffin[0] == "OFF" and buffin[1] == "ADJ"):
                        # wait 500 ms
                        time.sleep(0.5)
                        # query LTE CELL state
                        buffin = cmw.ask("SOURce:LTE:SIGN:CELL:STAT:ALL?")
                        # check TIMEOUT
                        if (time.time() - tstart) > TIMEOUT:
                            # x = win32api.MessageBox(0, "TIMEOUT occurs")
                            raise ValueError('TIMEOUT - LTE Signaling Turn OFF')
                else:
                    buffin = cmw.ask("SOUR:LTE:SIGN:CELL:STAT:ALL?").split(',')
                    while not (re.search('OFF', buffin[0], flags=re.IGNORECASE) and
                               re.search("ADJ", buffin[1], flags=re.IGNORECASE)):
                        # wait 500 ms
                        time.sleep(0.5)
                        # query LTE CELL state
                        buffin = cmw.ask("SOURce:LTE:SIGN:CELL:STAT:ALL?").split(',')
                        # check TIMEOUT
                        if (time.time() - tstart) > TIMEOUT:
                            # x = win32api.MessageBox(0, "TIMEOUT occurs")
                            raise ValueError('TIMEOUT - LTE Signaling Turn OFF')

                logger.info('LTE Cell - LTE CELL IS TURN OFF')
                logger.info('{0}'.format(132 * '-'))

                # Configure DAU
            if DAU_PRESENT:

                logger.info('{0}'.format(132 * '-'))
                logger.info('DAU Configure Services')

                for service in dau_service_l:

                    # Check DAU service
                    status = ''
                    if cmw_interface == "LAN":
                        status = cmw.ask("SOURce:DATA:CONT:{0}:STATe?".format(service))[0]
                    else:
                        status = cmw.ask("SOURce:DATA:CONT:{0}:STATe?".format(service))
                    logger.debug("Status of DAU : {0}".format(status))

                    # Turn OFF
                    if status != 'OFF':
                        cmw.write("SOURce:DATA:CONT:{0}:STATe OFF".format(service))

                        status = ''
                        while status != 'OFF':
                            time.sleep(0.150)
                            if cmw_interface == "LAN":
                                status = cmw.ask("SOURce:DATA:CONT:{0}:STATe?".format(service))[0]
                            else:
                                status = cmw.ask("SOURce:DATA:CONT:{0}:STATe?".format(service)).rstrip()

                        logger.debug('DAU service {0} is OFF'.format(service))

                # get IPV4 and IPV6 type
                if cmw_interface == "LAN":
                    buffipv4 = cmw.ask("CONFigure:DATA:CONTrol:IPVFour:ADDRess:TYPE?")[0]
                    buffipv6 = cmw.ask("CONFigure:DATA:CONTrol:IPVSix:ADDRess:TYPE?")[0]
                else:
                    buffipv4 = cmw.ask("CONFigure:DATA:CONTrol:IPVFour:ADDRess:TYPE?")
                    buffipv6 = cmw.ask("CONFigure:DATA:CONTrol:IPVSix:ADDRess:TYPE?")

                if 'external' in dau_ipv4_config or 'external' in dau_ipv6_config:

                    logger.debug('DAU -  Configure IP COnfig to External Network')

                    if ('external' in dau_ipv4_config and not 'DHCPv4' in buffipv4) or (
                            'external' in dau_ipv6_config and 'ACONf' not in buffipv6):
                        logger.debug('DAU -  Turn OFF DAU')

                        # turn DAU OFF
                        cmw.write("SOURce:DATA:CONTrol:STATe OFF")

                        status = ''
                        while status != 'OFF':
                            time.sleep(1.0)

                            status = cmw.ask("SOURce:DATA:CONTrol:STATe?")[0]

                    if 'external' in dau_ipv4_config and 'DHCPv4' not in buffipv4:
                        logger.debug('DAU -  Set IP Config IPV4 to DHCP')
                        cmw.ask("CONFigure:DATA:CONTrol:IPVFour:ADDRess:TYPE DHCPv4;*OPC?")

                    if 'external' in dau_ipv6_config and 'ACONf' not in buffipv6:
                        logger.debug('DAU -  Set IP Config IPV6 to DHCP')
                        cmw.ask("CONFigure:DATA:CONTrol:IPVSix:ADDRess:TYPE ACON;*OPC?")

                    # turn DAU ON
                    status = cmw.ask("SOURce:DATA:CONTrol:STATe?")[0]
                    if 'ON' not in status:

                        logger.debug('DAU -  Turn ON DAU')
                        cmw.write("SOURce:DATA:CONTrol:STATe ON")

                        status = ''
                        while status != 'ON':
                            time.sleep(1.0)
                            status = cmw.ask("SOURce:DATA:CONTrol:STATe?")[0]

                    logger.debug('DAU -  Check DAU IPConfig')

                    buffin = cmw.ask("SENSe:DATA:CONTrol:LAN:DAU:STATus?")[0]
                    if "CONN" not in buffin:
                        raise ValueError("DAU Not connected to external Ethernet Network.")

                else:

                    logger.debug('DAU -  Check  IP Config  Network')

                if cmw_interface == "LAN":
                    buffin = cmw.ask("CONFigure:DATA:CONTrol:IPVFour:ADDRess:TYPE?")[0]
                    if buffin != "AUT":
                        cmw.ask("CONFigure:DATA:CONTrol:IPVFour:ADDRess:TYPE AUT;*OPC?")

                    buffin = cmw.ask("CONFigure:DATA:CONTrol:IPVSix:ADDRess:TYPE?")[0]
                    if buffin != "AUTO":
                        cmw.ask("CONFigure:DATA:CONTrol:IPVSix:ADDRess:TYPE AUTO;*OPC?")

                    # DAU -  Configure DNS Service
                    logger.debug('DAU -  Configure DNS Service')
                    cmw.write(
                        'CONFigure:DATA:CONT:DNS:LOCal:ADD "epdg.epc.mnc{0}.mcc{1}.pub.3gppnetwork.org", "{2}"'.format(
                            MNC, MCC, epdg_ip_v4))
                    cmw.write(
                        'CONFigure:DATA:CONT:DNS:LOCal:ADD "epdg.epc.mnc{0}.mcc{1}.pub.3gppnetwork.org", "{2}"'.format(
                            MNC, MCC, epdg_ip_v6))

                # DAU IMS2 - Configure subscriber
                logger.debug('DAU IMS2 - Configure subscriber')
                cmw.write(
                    "CONF:DATA:CONT:IMS2:SUBScriber1:PRIVateid '{0}{1}{2}@msg.pc.t-mobile.com'".format(MCC, MNC, IMSI))
                cmw.write("CONF:DATA:CONT:IMS2:SUBScriber1:AUTHenticati:SCHeme AKA1")
                cmw.write("CONF:DATA:CONT:IMS2:SUBScriber1:AUTHenticati:ALGorithm XOR")
                cmw.write("CONF:DATA:CONT:IMS2:SUBScriber1:AUTHenticati:AMF '0x000'")
                cmw.write("CONF:DATA:CONT:IMS2:SUBScriber1:RESLength 128")
                cmw.write("CONF:DATA:CONT:IMS2:SUBScriber1:IPSec:ENABle ON")
                cmw.write("CONF:DATA:CONT:IMS2:SUBScriber1:IPSec:ALGorithm:INTegrity HMMD")
                cmw.write("CONF:DATA:CONT:IMS2:SUBScriber1:IPSec:ALGorithm:ENCRyption NOC")

                # Configure Virtual Subscriber
                logger.debug('DAU IMS2 - Configure Virtual Subscriber')
                cmw.write("CONF:DATA:CONT:IMS2:VIRTualsub1:BEHaviour ANSWER")
                cmw.write("CONF:DATA:CONT:IMS2:VIRTualsub1:SIGNalingtyp NOPRecondit")
                cmw.write("CONF:DATA:CONT:IMS2:VIRTualsub1:FORCemocall OFF")
                cmw.write("CONF:DATA:CONT:IMS2:VIRTualsub1:BEARer OFF")
                cmw.write("CONF:DATA:CONT:IMS2:VIRTualsub1:ADCodec:TYPE WIDeband")
                cmw.write('CONF:DATA:CONT:IMS2:VIRTualsub1:AMR:ALIGnment BANDwidtheff')
                cmw.write("CONF:DATA:CONT:IMS2:VIRTualsub1:AMR:CODec3:ENABle ON")

            else:
                logger.debug(" with  not Turning off the LTE signalling, DAU, Again ")

        # Configure LTE Signaling
        if 1:
            # time.sleep()
            logger.info('{0}'.format(132 * '-'))
            logger.info('Configure LTE Signaling')
            logger.info('LTE Cell - Set LTE Scenario To SISO')
            cmw.write(
                "ROUT:LTE:SIGN:SCEN:SCELl:FLEX SUA{0},{1},RX{2},{1},TX{2}".format(lte.sua, lte.rf_port, lte.converter))

            logger.info('LTE Cell - Set input and output path')
            logger.info(' including signal routing and Programming external attenuation.')
            #dl_att = 3.8
            #ul_att = 3.8
            cmw.write("CONF:LTE:SIGN:RFS:PCC:EATT:OUTP {0:0.2f}".format(float(DL_ATT)))  # 779 __<VAR
            cmw.write("CONF:LTE:SIGN:RFS:PCC:EATT:INP {0:0.2f}".format(float(UL_ATT)))  # 741
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            lte.mode = "TDD" if int(Band) in range(33, 42) else "FDD"
            logger.info('LTE Cell - Set LTE MODE to {0}'.format(lte.mode))
            cmw.write("CONF:LTE:SIGN:DMODe {0}".format(lte.mode))
            logger.info('{0}'.format(132 * '-'))

            if "TDD" in lte.mode:
                logger.info('{0}'.format(132 * '-'))
                logger.info('LTE - Set TDD ULDL to configuration 5')
                cmw.write("CONF:LTE:SIGN:CELL:ULDL 5")
                logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - Set System bandwidth')

            cmw.write("CONF:LTE:SIGN:CELL:BAND:DL B{0:03d}".format(int(10 * Channel_Bandwidth)))

            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - Set Operating Band and DL channel number')

            cmw.write("CONF:LTE:SIGN:BAND OB{0}".format(Band))

            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('Set Down Link Channel to: {0}'.format(float(DL_EARFCN)))
            cmw.write("CONF:LTE:SIGN:RFS:CHAN:DL {0}".format(DL_EARFCN))
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('Set UP Link Channel to: {0}'.format(UL_EARFCN))
            cmw.write("CONF:LTE:SIGN:RFS:CHAN:UL {0}".format(UL_EARFCN))
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - CELL ID')
            cmw.write("CONF:LTE:SIGN:CELL:PCID 0")
            logger.info('{0}'.format(132 * '-'))

            if cmw.sw_list.count('KS510') != 0:
                logger.info('{0}'.format(132 * '-'))
                logger.info('LTE Cell - CYCLE PREFIX')
                cmw.write("CONF:LTE:SIGN:CELL:CPR NORM")
                logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            cmw.write("CONF:LTE:SIGN:CELL:TIME:SATTach ON")
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - Define the RS EPRE level and the power of the DL channels and signals')
            logger.info(' Set RS EPRE level to {0:5.1f} dBm/FFT 15 KHz'.format(lte.rs_epre))
            cmw.write("CONF:LTE:SIGN:DL:RSEPre:LEVel {0:5.1f}".format(lte.rs_epre))
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('Advance PRACH settings')
            cmw.write("CONF:LTE:SIGN:UL:PCC:APPower:EASettings ON")
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - Set POWER CONTROL')
            cmw.write("CONF:LTE:SIGN:UL:TPC:SET MAXP")
            logger.info('Set LTE PUSCH Close LOOP Power to {0:5.1f} dBm'.format(float(pmx)))
            cmw.write("CONF:LTE:SIGN:UL:PUSCh:TPC:CLTPower {0:5.1f}".format(float(pmx)))
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE - Set Connection Type.')

            if DAU_PRESENT:
                logger.info('LTE - Set Connection Type to Data Connection.')
                cmw.write("CONF:LTE:SIGN:CONNection:CTYPe DAPP")

            else:
                logger.info('LTE - Set Connection Type to Test Mode Connection.')
                cmw.write("CONF:LTE:SIGN:CONNection:CTYPe TEST")
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - Set Network Parameters')

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - Specify MCC MNC, tracking area code and E-UTRAN cell ID.')
            cmw.write("CONF:LTE:SIGN:CELL:MCC {0}".format(MCC))
            cmw.write("CONF:LTE:SIGN:CELL:MNC:DIG THR")
            cmw.write("CONF:LTE:SIGN:CELL:MNC {0}".format(MNC))
            cmw.write("CONF:LTE:SIGN:CELL:TAC 1")
            cmw.write("CONF:LTE:SIGN:CELL:SEC:AUTH ON")
            cmw.write("CONF:LTE:SIGN:CELL:SEC:NAS ON")
            cmw.write("CONF:LTE:SIGN:CELL:SEC:AS ON")
            cmw.write("CONF:LTE:SIGN:CELL:SEC:IALG S3G")
            cmw.write("CONF:LTE:SIGN:CELL:SEC:MIL OFF")
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE Cell - Specify the User IMSI.')
            cmw.write("CONF:LTE:SIGN:CELL:UEID:IMSI '{0}{1}{2}'".format(MCC, MNC, IMSI))
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE - Set APN and CQI')
            cmw.write("CONF:LTE:SIGN:CONNection:APN '{0}'".format(lte.apn))
            cmw.write('CONF:LTE:SIGN:CONNection:QCI {0}'.format(lte.qci))
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            if IMS_FLAG:

                logger.info('NAS IMS Call supported')
                cmw.write("CONF:LTE:SIGN:CELL:NAS:EPSNetwork ON")
                cmw.write("CONF:LTE:SIGN:CELL:NAS:IMSVops SUPP")

            else:

                logger.info('NAS IMS Call not supported')
                cmw.write("CONF:LTE:SIGN:CELL:NAS:EPSNetwork OFF")
                cmw.write("CONF:LTE:SIGN:CELL:NAS:IMSVops NSUP")

            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE - Set LTE stack to Accept Multiple Bearer.')
            cmw.write("CONF:LTE:SIGN:CONNection:AMDBearer ON")
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE - Set LTE stack to for Redirection Inter LTE Band Handover')
            cmw.write("CONF:LTE:SIGN:CONNection:OBCHange RED")

            logger.info('LTE - Set LTE stack to for Redirection Inter LTE Channel Handover')
            cmw.write("CONF:LTE:SIGN:CONNection:FCHange RED")
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('LTE - Reduce PDCCH')
            cmw.write("CONF:LTE:SIGN:CONNection:PDCCh:RPDCch ON")
            logger.info('{0}'.format(132 * '-'))
            logger.info(' Configure RMC Test Mode Call. (User Defined Channel)')
            #cmw.write("CONF:LTE:SIGN:CONN:UET UDCH")
            #UL_TBS
            #cmw.write("CONF:LTE:SIGN:CONN:UDCH:DL {0},{1},{2}".format(DL_RB_Allocation, DL_RB_Start, DL_MODULATION))
            #cmw.write("CONF:LTE:SIGN:CONN:UDCH:UL {0},{1},{2}".format(UL_RB_Allocation, UL_RB_Start, UL_MODULATION))

            erreur = cmw.ask("SYST:ERR:ALL?")
            if erreur[0] != '0':
                raise ValueError(erreur)

            logger.info('{0}'.format(132 * '-'))

            # End Configure LTE CELL
        # *****************************************************************************
        # time.sleep(5)
        device.shell('settings put global airplane_mode_on 0')
        # NOTE: Some android versions require device.shell('su -c am broadcast -a android.intent.action.AIRPLANE_MODE')
        device.shell('am broadcast -a android.intent.action.AIRPLANE_MODE')

        # Start DAU services
        if DAU_PRESENT:
            logger.info('{0}'.format(132 * '-'))

            # Start DAU service
            logger.info(' Start DAU service(s)'.format(132 * '-'))

            for service in dau_service_l:

                logger.info('Start DAU {0} service'.format(service))
                status = cmw.ask("SOURce:DATA:CONT:{0}:STATe?".format(service))[0]

                if 'OFF' in status:
                    cmw.write("SOURce:DATA:CONT:{0}:STATe ON".format(service))

                    status = ''
                    while status != 'ON':
                        cmw.write("SOURce:DATA:CONT:{0}:STATe ON".format(service))
                        time.sleep(2)  # 0.150)
                        status = cmw.ask("SOURce:DATA:CONT:{0}:STATe?".format(service))[0]

                logger.info('DAU service {0} is ON'.format(service))

            erreur = cmw.ask("SYST:ERR:ALL?")
            if erreur[0] != '0':
                raise ValueError(erreur)

            logger.info('{0}'.format(132 * '-'))

            # Select FSW transducer file
            if test_environment.upper() != "PHASE1":
                pldir = os.path.join(os.getcwd(), "Pathloss")
                plfnam = pathloss_file[switch_path] + ".csv"
                plfile = os.path.join(pldir, plfnam)

            # Init Fsw
        if 1:
            logger.info('{0}'.format(132 * '-'))
            logger.info('Create a Socket Class object to remote the fsw')
            kwargs = {'timeout': fsw_socket_timeout, 'temchar': fsw_socket_termchar}

            # get instance for controlling CMW
            if fsw_interface == "LAN":
                fsw = c_socket(fsw_host, fsw_socket_port, **kwargs)
            else:
                fsw = c_gpib("FSW")

            # connect to FSW
            fsw.connect()
            fsw.timeout = 40.0
            logger.info('{0}'.format(132 * '-'))

            logger.info('{0}'.format(132 * '-'))
            logger.info('QUERY ID')

            if not hasattr(fsw, 'id'):
                setattr(fsw, 'id', 'FSW')
            if not hasattr(fsw, 'sn'):
                setattr(fsw, 'sn', '000000')
            if not hasattr(fsw, 'BASE_FW'):
                setattr(fsw, 'BASE_FW', 'FSW')

            buffin = fsw.ask("*IDN?")
            logger.debug(buffin)

            # get fsw options
            logger.info('{0}'.format(132 * '-'))
            logger.info('QUERY fsw Option')
            if not hasattr(fsw, 'opt'):
                setattr(fsw, 'opt', [])
            fsw.opt = fsw.ask("*OPT?")

            # fsw scpi commands start
            logger.info('{0}'.format(132 * '-'))
            logger.info('RESET FSW')
            #  Passing SCPI comands to FSW
            logger.debug(132 * '-')
            fsw.timeout = 30.0

            # fsw.write("*CLS")
            fsw.write("*RST")
            # fsw.write("SYST:PRES:ALL")
            time.sleep(0.5)

            fsw.write("INIT:CONT ON")
            logger.info('SCPI commands for FSW')
            logger.debug("run the Transducer file before Measurement")

            # if test_environment.upper() == "PHASE1":
            if (1):
                fsw.write("MMEM:LOAD:TFAC 'C:\FSW_pathloss_041321.csv'")
                fsw.write("SENSe:CORRection:TRANsducer:SELect 'FSW_pathloss_041321'")
            else:
                if not os.path.exists(plfile):
                    raise ValueError(("Error! file {} does not exists.".format(plfile)))
                logger.debug("Current transducer file: {}".format(plfile))
                fsw.write("MMEM:LOAD:TFAC {}".format(plfile))
                fsw.write("SENSe:CORRection:TRANsducer:SELect {}".format(pathloss_file[switch_path]))

            fsw.write("SENSe:CORRection:TRANsducer:STATe ON")

            # Frequency start and stop for measurement
            fsw.write("FREQ:STAR {0} MHz".format(int(freq_start_r1)))
            logger.debug("FREQ:STAR {0} MHz".format(int(freq_start_r1)))
            fsw.write("FREQ:STOP {0} MHz".format(int(freq_stop_r1)))
            logger.debug("FREQ:STOP {0} MHz".format(int(freq_stop_r1)))
            
            # Resolution Bandwidth and video Bandwidth
            fsw.write("BWIDth:RES {0} MHz".format(rbw))
            # Video Bandwidth)
            fsw.write("BWIDth:VID {0} MHz".format(vbw))
            logger.debug("FSW RBW={0}MHz VBW={1}MHz".format(rbw, vbw))
            
            # Ref_level and atten_ RF
            fsw.write("DISP:TRAC:Y:RLEV {}dBm".format(ref_level))
            fsw.write("ATT:AUTO OFF")
            fsw.write("INP:ATT {}dB".format(attn))
            
            # SWEEP POINTS
            fsw.write("SWE:POIN {}".format(sweep_point))

            # pre Amp
            fsw.write(":SENS:LIST:RANG1:INP:GAIN:STAT {0}".format(pre_amp))

            # limit power
            fsw.write("LIST:RANG1:LIM:STAR {0}".format(spec_limit))
            fsw.write("LIST:RANG1:LIM:STOP {0}".format(spec_limit))

            # setting detector type pos/RMS/.....
            fsw.write("CALC:MARKer1:FUNCtion:FMEasurement:DETector CRMS")
            
            # sweep count
            fsw.write("SWE:COUN {}".format(sweep_count))
            #fsw.write("INIT:CONT OFF")

            # Turn LTE Cell ON
        if 1:
            logger.info('{0}'.format(132 * '-'))

            def lte_turn_on():
                logger.info('TURN LTE CELL ON')
                cmw.write("SOUR:LTE:SIGN:CELL:STATe ON")
                TIMEOUT = 60.0
                tstart = time.time()
                buffin = None

                if cmw_interface == "LAN":
                    buffin = cmw.ask("SOUR:LTE:SIGN:CELL:STAT:ALL?")
                else:
                    buffin = cmw.ask("SOUR:LTE:SIGN:CELL:STAT:ALL?").split(',')

                if cmw_interface == "LAN":
                    while not (buffin[0] == 'ON' and buffin[1] == "ADJ"):
                        buffin = cmw.ask("SOURce:LTE:SIGN:CELL:STAT:ALL?")
                        time.sleep(0.5)
                        if (time.time() - tstart) > TIMEOUT:
                            # x = win32api.MessageBox(0, "TIMEOUT occurs")
                            raise ValueError('TIMEOUT - LTE Signaling Turn OFF')
                else:
                    while not (re.search('ON', buffin[0], flags=re.IGNORECASE) and
                               re.search("ADJ", buffin[1], flags=re.IGNORECASE)):
                        buffin = cmw.ask("SOURce:LTE:SIGN:CELL:STAT:ALL?").split(',')
                        # logger.debug("LTE Signaling = '{}'".format(buffin))
                        time.sleep(0.5)
                        if (time.time() - tstart) > TIMEOUT:
                            raise ValueError('TIMEOUT - LTE Signaling Turn OFF')

                logger.info('LTE CELL READY')
                logger.info('{0}'.format(132 * '-'))
            time.sleep(2)
            lte_turn_on()
            # Check Default Bearers
            TIMEOUT = 30
            # changed
            tstart = time.time()

            if cmw_interface == "LAN":
                buffin = cmw.ask('CATalog:LTE:SIGN:CONNection:DEFBearer?')[0]  ##{ ["5 (Test Network)","6 (ims)"]}
                logger.debug("Bearers in buffin {0}".format(buffin))
                state = cmw.ask('SENSE:LTE:SIGN:RRCState?')[0]
                logger.debug("cmw RRC state is : {0}".format(state))
            else:
                buffin = cmw.ask('CATalog:LTE:SIGN:CONNection:DEFBearer?')
                logger.debug("Bearers in buffin {0}".format(buffin))
                state = cmw.ask('SENSE:LTE:SIGN:RRCState?')
                logger.debug("cmw RRC state is : {0}".format(state))

            device.shell('ip addr >DUT_ip.txt')
            device.shell('devices')
            while not "5 (Test Network)" in buffin:
                if cmw_interface == "LAN":
                    buffin = cmw.ask('CATalog:LTE:SIGN:CONNection:DEFBearer?')[0]
                else:
                    buffin = cmw.ask('CATalog:LTE:SIGN:CONNection:DEFBearer?')

                time.sleep(1)
                if (time.time() - tstart) > TIMEOUT:
                    logger.info('TIMEOUT - Default Bearers not present')
                    # Toggle airplane mode on -> off
                    # device.shell('root')
                    # time.sleep(2.0)
                    device.shell('settings put global airplane_mode_on 1')
                    # NOTE: Some android versions require device.shell('su -c am broadcast -a android.intent.action.AIRPLANE_MODE')
                    device.shell('am broadcast -a android.intent.action.AIRPLANE_MODE')
                    time.sleep(0.5)
                    device.shell('settings put global airplane_mode_on 0')
                    device.shell('am broadcast -a android.intent.action.AIRPLANE_MODE')
                    time.sleep(1.0)
                    # wait for dut to register
                    logger.info('waiting for dut to register. airplane mode on -> off')
                    start_register_time = time.time()
                    registered = False
                    while not registered and time.time() < start_register_time + REGISTER_TIMEOUT:
                        register_state = cmw.ask('FETCH:LTE:SIGN:PSW:STATe?')[0]
                        rrc_state = cmw.ask('SENSE:LTE:SIGN:RRCState?')[0]
                        logger.info('connection state : %s : %s', register_state, rrc_state)

                        if rrc_state == 'CONN':
                            registered = True
                        else:
                            # time.sleep(1.0)
                            device.shell('reboot')
                            lte_turn_on()
                            time.sleep(90)
                            device.shell('root')
                            time.sleep(10)
                    if not registered:
                        logger.error('dut did not register successfully. terminating.')
                        break
                        # sys.exit(-3)

                    logger.info('dut did register successfully.')
            logger.info('Default Bearer detected')

            # device.shell('ifconfig')
            ip_addr = device.shell('netcfg')
            logger.debug(ip_addr)

            # device.shell('ip addr')

            for i in buffin:
                if i == "5 (Test Network)":
                    result = "PASS"
                else:
                    result = "FAIL"

            # Enabling TX Measurement
            cmw.write("CONF:LTE:SIGN:CONN:UET UDCH")
            # UL_TBS
            # cmw.write("CONF:LTE:SIGN:CONN:UDCH:DL {0},{1},{2},{3}".format(DL_RB_Allocation, DL_RB_Start, DL_MODULATION,DL_TBS))
            # cmw.write("CONF:LTE:SIGN:CONN:UDCH:UL {0},{1},{2},{3}".format(UL_RB_Allocation, UL_RB_Start, UL_MODULATION, UL_TBS))
            tx_mev = True
            if tx_mev:
                logger.debug("LTE TX Multi-Evaluation TEST")
                logger.debug("INFO:*** Route LTE MEV in Combine Signal Path")

                cmw.write("ROUTe:LTE:MEAS:SCENario:CSP 'LTE Sig1'")
                cmw.write('ABORt:LTE:MEAS:MEValuation')
                cmw.write('INIT:LTE:MEAS:MEValuation')
                time.sleep(0.25)

                TX_buffin = cmw.ask('FETCh:LTE:MEAS:MEValuation:LIST:SEGMent1:MODulation:DALLocation?')
                logger.debug('TX Meas is: {0}'.format(TX_buffin))
                TX_demodulation = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:DMODulation?")
                logger.debug('TX demodulation is: {0}'.format(TX_demodulation))
                TX_DCHType = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:DCHType?")
                logger.debug('TX DCHType is: {0}'.format(TX_DCHType))

                if cmw_interface == "LAN":
                    Txmeas = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:CURRent?")
                else:
                    Txmeas = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:CURRent?")[:-1].split(',')

                logger.debug('TX Meas is: {0}'.format(Txmeas))
                res = ['1_Reliability', '2_OutOfTol', '3_EVM_RMSlow', '4_EVM_RMShigh', '5_EVMpeakLow', '6_EVMpeakHigh',
                       '7_MErr_RMSlow', '8_MErr_RMShigh', '9_MErrPeakLow', '10_MErrPeakHigh', '11_PErr_RMSlow',
                       '12_PErr_RMSh', '13_PErrPeakLow', '14_PErrPeakHigh', '15_IQoffset', '16_FreqError',
                       '17_TimingError', '18_TXpower', '19_PeakPower', '20_RBpower', '21_EVM_DMRSl', '22_EVM_DMRSh',
                       '23_MErr_DMRSl', '24_MErr_DMRSh', '25_PErr_DMRS', '26_PErr_DMRSh', '27_GainImbal',
                       '28_QuadError', '29_EVM_SRS']
                dict1 = {}

                for i_meas, meas in enumerate(Txmeas):
                    if 'E+' in meas or 'E-' in meas:
                        Txmeas[i_meas] = '{0:8.3f}'.format(float(meas))
                    logger.debug('{0}:{1:>8}'.format(res[i_meas], Txmeas[i_meas]))
                    dict2 = {'{0}'.format(res[i_meas]): (Txmeas[i_meas])}
                    dict1.update(dict2)

                logger.debug("dict1 is here", dict1)


                def extractDigits(lst):
                    res = []
                    for el in lst:
                        sub = el.split(', ')
                        res.append(sub)
                    return (res)


                # Driver code
                lst = dict1.values()
                lst2 = extractDigits(lst)
                logger.debug(lst2)
                Dict_dataframe = dict(zip(res, lst2))
                Tx_MultiEval_Res = {}
                for key, v in Dict_dataframe.items():
                    if key == '18_TXpower' or key == '4_EVM_RMShigh' or key == '15_IQoffset' or key == '16_FreqError':
                        Tx_MultiEval_Res[key] = v
                logger.debug("multivaluation result is ", Tx_MultiEval_Res)

                DUTtxpower = Tx_MultiEval_Res['18_TXpower']
                Tx_P = str(DUTtxpower[0])

                logger.debug(132 * '_')
                logger.debug(132 * '_')
                logger.debug("DUT TX power is : {0}".format(Tx_MultiEval_Res.get("18_TXpower")))
                logger.debug(132 * '_')
                logger.info('{0}'.format(132 * '-'))

                """
                if cmw_interface == "LAN":
                    Tx_multi_avg_result = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:AVERage?")
                else:
                    Tx_multi_avg_result = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:AVERage?").split(',')

                res = ['1_Reliability', '2_OutOfTol', '3_EVM_RMSlow', '4_EVM_RMShigh', '5_EVMpeakLow', '6_EVMpeakHigh',
                       '7_MErr_RMSlow', '8_MErr_RMShigh', '9_MErrPeakLow', '10_MErrPeakHigh', '11_PErr_RMSlow',
                       '12_PErr_RMSh', '13_PErrPeakLow', '14_PErrPeakHigh', '15_IQoffset', '16_FreqError',
                       '17_TimingError', '18_TXpower', '19_PeakPower', '20_RBpower', '21_EVM_DMRSl', '22_EVM_DMRSh',
                       '23_MErr_DMRSl', '24_MErr_DMRSh', '25_PErr_DMRS', '26_PErr_DMRSh', '27_GainImbal',
                       '28_QuadError', '29_EVM_SRS']

                for i_meas, meas in enumerate(Tx_multi_avg_result):
                    if 'E+' in meas or 'E-' in meas:
                        Tx_multi_avg_result[i_meas] = '{0:8.3f}'.format(float(meas))
                    logger.debug('{0}:{1:>8}'.format(res[i_meas], Tx_multi_avg_result[i_meas]))
                    logger.debug("Tx multievaluation average:{0}".format(Tx_multi_avg_result))

                if cmw_interface == "LAN":
                    Tx_multi_extreme_result = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:EXTReme?")
                else:
                    Tx_multi_extreme_result = cmw.ask("FETCh:LTE:MEAS:MEValuation:MODulation:EXTReme?").split(',')

                res1 = ['1_Reliability', '2_OutOfTol', '3_EVM_RMSlow', '4_EVM_RMShigh', '5_EVMpeakLow', '6_EVMpeakHigh',
                        '7_MErr_RMSlow', '8_MErr_RMShigh', '9_MErrPeakLow', '10_MErrPeakHigh', '11_PErr_RMSlow',
                        '12_PErr_RMSh', '13_PErrPeakLow', '14_PErrPeakHigh', '15_IQoffset', '16_FreqError',
                        '17_TimingError', '18_TXpowerMin', '19_TXpowerMax', '20_PeakPowerMin', '21_PeakPowerMax',
                        '22_RBpowerMin', '23_RBpowerMax', '24_EVM_DMRSl', '25_EVM_DMRSh', '26_MErr_DMRSl',
                        '27_MErr_DMRSh', '28_PErr_DMRS', '29_PErr_DMRSh', '30_GainImbal', '31_QuadError', '32_EVM_SRS']

                for i_meas, meas in enumerate(Tx_multi_extreme_result):
                    if 'E+' in meas or 'E-' in meas:
                        Tx_multi_extreme_result[i_meas] = '{0:8.3f}'.format(float(meas))
                    logger.debug('{0}:{1:>8}'.format(res1[i_meas], Tx_multi_extreme_result[i_meas]))
                logger.debug("Tx multievaluation extreme:{0}".format(Tx_multi_extreme_result))
                """
                
                """
                Tx_multi_SDEViation_result = cmw.read("READ:LTE:MEAS<i>:MEValuation:MODulation:SDEViation?")
                res = ['1_Reliability', '2_OutOfTol', '3_EVM_RMSlow', '4_EVM_RMShigh', '5_EVMpeakLow', '6_EVMpeakHigh',
                       '7_MErr_RMSlow', '8_MErr_RMShigh', '9_MErrPeakLow', '10_MErrPeakHigh', '11_PErr_RMSlow',
                       '12_PErr_RMSh', '13_PErrPeakLow', '14_PErrPeakHigh', '15_IQoffset', '16_FreqError',
                       '17_TimingError', '18_TXpower', '19_PeakPower', '20_RBpower', '21_EVM_DMRSl', '22_EVM_DMRSh',
                       '23_MErr_DMRSl', '24_MErr_DMRSh', '25_PErr_DMRS', '26_PErr_DMRSh', '27_GainImbal',
                       '28_QuadError', '29_EVM_SRS']
                for i_meas, meas in enumerate(Tx_multi_SDEViation_result):
                    if 'E+' in meas or 'E-' in meas:
                        Tx_multi_SDEViation_result[i_meas] = '{0:8.3f}'.format(float(meas))
                    logger.debug('{0}:{1:>8}'.format(res[i_meas], Tx_multi_SDEViation_result[i_meas]))
                logger.debug("Tx_multi_SDEViation_result:{0}".format(Tx_multi_SDEViation_result))
                """

            # Switch path for FSW measurement
            if sw_box_interface == "GPIB":
                swbox.set_sw_box(switch_path)

            # Set peak search
            fsw.write("CALC:PSE:MARG 100")
            fsw.write("CALC:PSE:PSH ON")
            fsw.write("CALC:PSE:SUBR 1")
            fsw.write("CALC:MARK:MAX:AUTO ON")
            
            if (freq_start_r1 > 3000):
                time.sleep(3)  # 6)
            else:
                time.sleep(1)  # 3)
            
            if fsw_interface == "LAN":
                peak_y = float(fsw.ask("CALC:MARK1:Y?")[0])
                peak_x = float(fsw.ask("CALC:MARK1:X?")[0])
            else:
                peak_y = float(fsw.ask("CALC:MARK1:Y?"))
                time.sleep(0.5)
                peak_x = float(fsw.ask("CALC:MARK1:X?"))
                time.sleep(0.5)
            
            list_of_Freq = [round(peak_x/1000), peak_y, peak_y - spec_limit]
            logger.debug("list_of_freq: {}".format(list_of_Freq))
            
            # FSW instrument screeshot
            fsw.write("FORM:DEXP:FORM CSV")
            fsw.write("FORM:DEXP:DSEP POIN")
            fsw.write("FORM:DEXP:FORM CSV")
            fsw.write("FORM:DEXP:HEAD ON")
            fsw.write("FORM:DEXP:TRAC SING")
            fsw.write("DISP:WIND1:TRAC:PRES MCM")

            # Directory path of csv
            fsw.write("MMEM:STOR1:TRAC 1,'C:\\R_S\\instr\\user\\fsw_Results_{0}.CSV'".format(row_count))
            logger.debug(len(list_of_Freq))
            logger.debug(list_of_Freq)

            if len(list_of_Freq) == 3:

                # Printing the Spurious emissions verdict
                logger.debug(list_of_Freq)

                n = 3
                m = np.array(list_of_Freq)
                y = m.astype(float)
                z = y.tolist()
                Meas_range = [z[i:i + n] for i in range(0, len(z), n)]

                def Extract(Meas_range):
                    Freq = list(map(itemgetter(0), Meas_range))
                    Power_abs = list(map(itemgetter(1), Meas_range))
                    Limit_p = list(map(itemgetter(2), Meas_range))
                    return Freq, Power_abs, Limit_p

                Meas = Extract(Meas_range)
                for i in Meas:
                    logger.debug(i)

                """
                abs_fre = [Meas[1][0]]
                logger.debug ("abs power msw value are {0}".format(Meas[1][0]))
                # Measured values of ranges 1,2,3 should be less than -36 dbm
                logger.debug(abs_fre)
                logger.debug(len(abs_fre))

                if abs_fre <= float(spec_limit):
                    logger.debug("abs _power {0}".format(abs_fre))
                    verdict.extend(['PASS'])
                else:
                    logger.debug("abs _power {0}".format(abs_fre))
                    verdict.extend(['FAIL'])

                # range 4 value
                if abs_fre[3] <= float(min) :
                    verdict.extend(['PASS'])
                    logger.debug("abs _power {0}".format(abs_fre[3]))

                else:
                    verdict.extend(['FAIL'])
                if abs_fre[4] <= float(min) :
                    verdict.extend(['PASS'])
                    logger.debug("abs _power {0}".format(abs_fre[4]))
                else:
                    verdict.extend(['FAIL'])
                """

                # for max values of 5 ranges verdict
                ranges = [1]
                freq_ranges = [Meas[0][0]]
                abs_power = [Meas[1][0]]
                limit = [Meas[2][0]]
                band_info = [Band]
                DL_freq = [DL_EARFCN]
                # initialization of result list
                # ref_power = [-36, -36, -36, -30, -30]
                #margin_l = []
                #zip_object = zip(abs_power, limit)
                #for list1_i, list2_i in zip_object:
                #    margin_l.append(list1_i - list2_i)

                # append each difference to list
                #logger.debug(margin_l)  # Ref_level
                #margin = []  # Diff 0f -36-(-limit)
                #zip_object = zip(margin_l, limit)
                #for list1_j, list2_j in zip_object:
                #    margin.append(list2_j - list1_j)
                
                margin = []  # Diff 0f -36-(-limit)
                margin.append(spec_limit - abs_power[0])

                # append each difference to list
                logger.debug("margin={}".format(margin))

                # min = -30.0
                # max = -36.0
                verdict = []
                # summary =[]
                if margin[0] > 0:
                    # summary = 'PASS'
                    verdict = 'PASS'
                else:
                    # summary = 'FAIL'
                    verdict = 'FAIL'

                # verdict list
                logger.debug(Meas)
                logger.debug(132 * '-')
                logger.debug("Measurement Ranges for FSW spectrum analyzer")
                logger.debug(132 * '-')
                logger.debug(" Range  |  Freq  kHZ| PowerAbs(-dbm) |  limit (-db) | Margin(-dbm) | Verdict  ")
                logger.debug(
                    " Range | {0:.05f}     | {1:.02f} |  {2:.02f} ||  {3:.02f} |{4}  ".format(Meas[0][0] / 1000,
                                                                                               Meas[1][0],
                                                                                               Meas[2][0], margin[0],
                                                                                               verdict[0]))

                # logger.debug(" Test Summary for Frequency Measurement for all ranges : {0}  ".format(summary))
                logger.debug(" Test Summary for Frequency Measurement for all ranges : {0}  ".format(verdict))
                logger.debug(132 * '-')
                # writing to file Fsw measuremnt

                # dictionary of lists

                dict_f = {'Range': ranges, 'Band': Band, 'DL_freq': DL_EARFCN, 'RBW': rbw, 'frequency': freq_ranges,
                          'abs_power': abs_power, 'limit_pwr': limit, 'margin(-dbm)': margin}
                df = pd.DataFrame(dict_f)
                timestr = time.strftime("%Y%m%d-%H%M%S")
                # saving the dataframe
                logger.debug(132 * '_')
                logger.debug(mydir_d)
                logger.debug(132 * '_')
                # df.to_csv('{3}\\fsw_logsresult_{0}_band{1}_{2}.csv'.format(Channel_Bandwidth, TEST_BAND, timestr, mydir), index=False, mode='a')

            else:
                logger.debug(" fsw intialization fail or either FSW spurious msmt  values are 1 or 0 ")
                logger.debug("FSW scpi command fsw.ask(TRAC:DATA? SPURIOUS) returns either value 1  or 0 in msmt list ")

        # End Init FSW
        erreur = cmw.ask("SYST:ERR:ALL?")
        logger.debug(
            "Band: 	\t Bandwidth : (MHz) \t	DL Frequency:  (MHz)	\tDL RB Allocation: 	\t DL RB Start: 	\t float(Power Level)")
        logger.debug(
            " {0}	 \t\t  :{1} (MHz) \t\t	: {2} (MHz)\t	\t\t: {3}	   \t\t\t : {4}	\t\t\t  {5}  ".format(
                Band, Channel_Bandwidth, DL_EARFCN, DL_RB_Allocation, DL_RB_Start, pmx))
        # Storing the output in json
        rth = RsInstrument('TCPIP::{}::INSTR'.format(screenshot_host_ip))
        rm = pyvisa.ResourceManager()
        instr = rm.open_resource('TCPIP::{}::INSTR'.format(screenshot_host_ip))  # replace by your IP-address
        instr.timeout = 3000
        instr.write('INIT:CONT OFF')
        # truns on color printing
        instr.write('HCOP:DEV:COL ON')
        # select file format
        # (WMF | GDI | EWMF | BMP | PNG | JPEG | JPG | PDF | SVG | DOC | RTF)
        instr.write('HCOP:DEV:LANG PNG')

        # set print to file
        instr.write('HCOP:DEST "MMEM"')
        timestr1_cmw = time.strftime("%Y%m%d-%H%M%S")
        logger.debug(132 * '_')
        logger.debug(mydir_d)
        logger.debug(132 * '_')
        filePathInstr = r"/temp/spurious_emi{0}_{1}_{2}.png".format(Band, Channel_Bandwidth, timestr1_cmw)  # size of 18kb
        filePathPc = r"{4}\\{0}_spurious_emi_{1}_{2}_{3}.png".format(row_count, Band, Channel_Bandwidth, timestr1_cmw, mydir)  # no data
        logger.debug(filePathPc)
        #screenshot = filePathPc.split("\\")[4]
        # file path on instrument
        instr.write('MMEM:NAME "{}"'.format(filePathInstr))
        # create screenshot
        instr.write("HCOP:IMM")
        # ask for file data from instrument
        fileData = instr.query_binary_values('MMEM:DATA? "{}"'.format(filePathInstr), datatype='s')[0]

        rth.write_str(f"MMEM:NAME '{filePathInstr}'")
        rth.write_str_with_opc("HCOP:IMM")
        rth.read_file_from_instrument_to_pc(filePathInstr, filePathPc)
        print(instr.query('SYST:ERR?'))
        instr.close()
        Build_info = "facebook/mos/mos:8.1.0/OPM1.171019.026/15771000000000000"
        logger.debug(filePathPc)
        logger.debug(type(filePathPc))
        screenshot = filePathPc.split("\\")[4]
        logger.debug(screenshot)
        Margin_Range1 = margin[0]

        Verdict_ctf = []
        if Tx_P == 'INV' or Tx_P == 'NAV':
            Verdict_ctf = 'INCONCLUSIVE'
        else:
            if margin[0] >= 3:
                Verdict_ctf = 'PASS'
            elif margin[0] < 0:
                Verdict_ctf = 'FAIL'
            else:
                Verdict_ctf = 'MARGINAL'
        #header_list =
        lte_tx_result = {'VERDICT': [Verdict_ctf],'TEMP':[Temp], 'VOLTAGE':[Voltage], 'TECH':[Tech], 'ANTENA':[Antena], 'NS_FILTER':[NS_Filter], 'PMAX':[pmx],
       'BAND':[Band], 'UL EARFCN':[UL_EARFCN], 'DL EARFCN':[DL_EARFCN], 'CHANNEL BANDWIDTH':[Channel_Bandwidth],
       'UL RB Allocation':[UL_RB_Allocation], 'UL RB Start':[UL_RB_Start], 'UL MODULATION':[UL_MODULATION], 'UL_TBS':[UL_TBS],
       'DL RB Allocation':[DL_RB_Allocation], 'DL RB START':[DL_RB_Start], 'DL MODULATION':[DL_MODULATION], 'RB CONNECTION':[DL_TBS],
       'START_FREQ':[freq_start_r1], 'STOP_FREQ':[freq_stop_r1], 'DELFOB':[DelFOB], 'SWITCH':[switch_path], 'REF_LEVEL':[ref_level], 'ATTN':[attn],
       'RBW':[rbw], 'VBW':[vbw], 'PRE_AMP':[pre_amp], 'TRACE':[Trace], 'DETECTOR':[Detector], 'SWEEP_TIME':[sweep_time],
       'SWEEP_POINT':[sweep_point], 'SWEEP_COUNT':[sweep_count],'TX PWR(dB)': [Tx_P],'SPUR PEAK FRQ RANGE1 (KHz)': [freq_ranges[0] / 1000],'SPUR ABS_PWR RANGE1(dBm)': [abs_power[0]], 'SPUR_LIMIT_RANGE1(dBm)': [margin[0]],'MARGIN_RANGE1(dBm)': [margin[0]], 'SCREENSHOT': [screenshot]}
        # lte_tx_result = {'VERDICT': [Verdict_ctf], "BAND": [Band], "UPLINK FREQUENCY": [UL_EARFCN],
        #                  'C_BW(MHz)': [Channel_Bandwidth],
        #                  'UL RB': [UL_RB_Allocation], 'UL RB START': [UL_RB_Start], 'TX PWR(-db)': [Tx_P],
        #                  'SPUR PEAK FRQ RANGE1 (KHz)': [freq_ranges[0] / 1000],
        #                  'SPUR ABS_PWR RANGE1(-dbm)': [abs_power[0]], 'SPUR_LIMIT_RANGE1(-db)': [margin_l[0]],
        #                  'MARGIN_RANGE1(-dbm)': [margin[0]], 'SCREENSHOT': [screenshot]}

        output = pd.DataFrame(lte_tx_result)

        timestr_cmw = time.strftime("%Y%m%d-%H%M%S")
        # saving the dataframe
        logger.debug(132 * '_')
        logger.debug(mydir)
        logger.debug(132 * '_')
        output.to_csv('{1}\\CMW_logsresult_{0}.csv'.format(timestr_cmw, mydir), index=False, mode='a')
        #
        """
        if erreur[0] != '0':
            raise ValueError(erreur)
        logger.info('{0}'.format(132 * '-'))
        """

    # catch error
    except Exception as Err:
        logger.error(Err)
        logging.traceback.print_exc()
        break
    finally:
        # close CMW class
        if cmw is not None:
            cmw.write('&GTL')
            cmw.close()
            cmw = None
    row_count = row_count + 1

    logger.debug("end of {0} th loop".format(int(row_count)))

# Storing all csv file in to single file to conclude the results summary for all runs
timestr_out = time.strftime("%Y%m%d")
logger.debug(132 * '_')
logger.debug(mydir_d)
logger.debug(132 * '_')
# defining glob function to aggregate the CMW, FSW CSV files
logger.debug(132 * '_')
# logger.debug("my destination folder is {0}".format(dest_d))
cmw_csv_files = glob("{1}\\CMW_logsresult_*.csv".format(timestr_out, mydir))
# creating pandas data frame dict for cmw csv files
df = pd.concat((pd.read_csv(f, header=0) for f in cmw_csv_files))
timestr_f = time.strftime("%Y%m%d-%H%M%S")
LTE_TX_SPURIOUS_EMISSIONS_OUTPUT = 'LTE_TX_SPURIOUS_EMISSIONS_OUTPUT_{0}'.format(
    timestr_f)  # CMW_VERDICT_20210221-221813
# writing csv files
df.to_csv("{1}\\{0}.csv".format(LTE_TX_SPURIOUS_EMISSIONS_OUTPUT, mydir), index=False)

# writing on to csv files content to json file
""""
FSW_verdict = 'FSW_VERDICT_{0}'.format(timestr_f)
fsw_csv_files = glob("{1}\\fsw_*.csv".format(timestr_out, mydir))
df_fsw = pd.concat((pd.read_csv(f, header=0 ) for f in fsw_csv_files))
df_fsw.to_csv("{1}\\{0}.csv".format(FSW_verdict, mydir), index=False)
"""
# converting csv file to json
# moving csv files and json files to results_run folder
# Source path
if os.path.isdir(dest_d):
    try:
        shutil.rmtree(dest_d)
        logger.debug("file is deleted ")
    except OSError as e:
        logger.debug("Error: %s : %s" % (dest_d, e.strerror))

shutil.copytree(mydir, dest_d, dirs_exist_ok=True)

for fname in os.listdir(dest):
    if fname.startswith("CMW"):
        os.remove(os.path.join(dest, fname))
#screenshot_folder = "{0}\SCREENSHOT".format(dest)


# Source path
source = "{0}\\{1}".format(os.getcwd(), lte_tx_input_file)
# Destination path
destination = dest_d
# Copy the content of
# source to destination
shutil.copy(source, destination)
# clearing Output file
#sort = scr_shot
#source = dest_d
scr_path = os.path.join(dest, "screenshot")
scr_shot = os.mkdir(scr_path)
logger.debug(scr_path)
sort = scr_path

for f in os.listdir(dest_d):
    if f.endswith(".png"):
        shutil.move(os.path.join(dest_d, f), sort)
# Moving .log file to final
# for f1 in os.listdir(dest_d.split('LTE_TX_harmonics_FULL_RUN_LOG')[0]):
#     if f1.endswith(".log"):
#         shutil.move(os.path.join(dest_d.split('LTE_TX_harmonics_FULL_RUN_LOG')[0], f1), dest)
# """
# converting fsw verdict csv file to json
"""
df_f = pd.read_csv (r'{0}\\{1}.csv'.format(mydir, FSW_verdict))
#df_j.to_json (r'{0}\\ctf_j.json'.format(mydir_d))
df.to_json (r'{0}\\ctf_fsw.json'.format(mydir), orient='split')
# printing the csv contents in run log .py for all variants
"""

"""
This script:
1. collects the test case csv filenames into a list
2. reads each csv file into a dataframe
3. convert Series to a dictionary, adds data sources to json
4. adds specified charts
5. save data series to json file for visualization
# Sample usage
    cv = CtfVisualization(path=<path to test case files directory>)
    cv.create_json()
Note: Remove any non test case csv files from directory before running to prevent conflict 
or extraneous data from being included in the visualization.
"""


class CtfVisualization:
    """Interface to create the visualization data json file from csv results files"""
    _description: str  # not currently used
    _name: str
    _path: str
    _data_sources: list

    def __init__(self, description: str = "Visualize test case csv files",
                 name: str = "LTE_TX_SPURIOUS_EMISSIONS_OUTPUT.json", path: str = ""):
        self._description = description
        self._name = name
        self._path = path
        self._data_sources = []

    def create_json(self):
        # CTF create data blob
        ctf_json = CtfJsonData(name=self._name, path=self._path)

        all_csv_files = glob(os.path.join(self._path, "LTE_TX_SPURIOUS_EMISSIONS_OUTPUT_*.csv"))
        for fn in all_csv_files:
            df = pd.read_csv(fn, sep=',', na_filter=False)
            # CTF add data source and table
            ctf_data_source = os.path.basename(fn)
            self._data_sources.append(ctf_data_source)
            ctf_json.add_data_source(ctf_data_source)
            ctf_columns = (', '.join(df))
            ctf_json.add_table(title=ctf_data_source, columns=ctf_columns, data_source_list=ctf_data_source)

            # CTF add data to source
            df_dict_ctf = df.T.to_dict()
            for row in df_dict_ctf:
                ctf_json.add_data_to_source(ctf_data_source, [df_dict_ctf[row]])

        # CTF dump to JSON file
        ctf_json.save_data()


# Sample usage
if __name__ == "__main__":
    cv = CtfVisualization(path=dest_d)
    cv.create_json()

#shutil.move(logfilename,dest_d)

# THE END
